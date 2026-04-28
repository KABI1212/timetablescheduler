from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Optional, Set, Tuple
from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

from scheduler import Subject, Teacher, Room, TimetableScheduler


app = FastAPI(title="Smart Classroom Timetable Scheduler", version="1.0.0")


class SubjectInput(BaseModel):
    subject_name: str
    theory_hours_per_week: int
    lab_hours_per_week: int
    teacher_assigned: str


class TeacherInput(BaseModel):
    teacher_name: str
    subjects_they_can_teach: List[str]
    availability: Optional[List[Tuple[int, int]]] = None  # (day, period)


class RoomInput(BaseModel):
    room_id: str
    room_type: str  # classroom | lab


class GenerateRequest(BaseModel):
    subjects_by_section: Dict[str, List[SubjectInput]]
    teachers: List[TeacherInput]
    rooms: List[RoomInput]
    days_per_week: int = 5
    periods_per_day: int = 8
    optimize: bool = True


class ExportExcelRequest(BaseModel):
    timetable: List[Dict[str, object]]


@app.get("/")
def home() -> str:
    return "Smart Classroom Timetable Scheduler API is running."


@app.get("/health")
def health() -> Dict[str, str]:
    return {
        "status": "ok",
        "scheduler": "simulated_annealing",
        "version": "1.0.0"
    }


@app.post("/generate")
def generate(req: GenerateRequest) -> Dict[str, object]:
    teachers: Dict[str, Teacher] = {}
    for t in req.teachers:
        availability_set: Optional[Set[Tuple[int, int]]] = None
        if t.availability:
            availability_set = set((int(d), int(p)) for d, p in t.availability)
        teachers[t.teacher_name] = Teacher(
            teacher_name=t.teacher_name,
            subjects_they_can_teach=set(t.subjects_they_can_teach),
            availability=availability_set
        )

    rooms = [Room(r.room_id, r.room_type) for r in req.rooms]

    subjects_by_section: Dict[str, List[Subject]] = {}
    for section, subs in req.subjects_by_section.items():
        subjects_by_section[section] = [
            Subject(
                subject_name=s.subject_name,
                theory_hours_per_week=s.theory_hours_per_week,
                lab_hours_per_week=s.lab_hours_per_week,
                teacher_assigned=s.teacher_assigned
            )
            for s in subs
        ]

    scheduler = TimetableScheduler(
        subjects_by_section=subjects_by_section,
        teachers=teachers,
        rooms=rooms,
        days_per_week=req.days_per_week,
        periods_per_day=req.periods_per_day
    )

    schedule = scheduler.generate_timetable()
    if req.optimize:
        scheduler.optimize_with_ai()

    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][:req.days_per_week]
    rows = scheduler.to_rows(day_names)

    return {
        "timetable": rows,
        "grid": scheduler.to_grid(),
        "classroom_utilization": scheduler.classroom_utilization_report(),
        "teacher_workload": scheduler.teacher_workload_analysis(),
        "rejected_slots": scheduler.get_rejected_slots(day_names),
        "total_sessions": len(schedule)
    }


@app.post("/export-excel")
def export_excel(payload: ExportExcelRequest) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timetable"

    headers = ["Day", "Period", "Section", "Subject", "Teacher", "Room", "Type"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in payload.timetable:
        sheet.append([row.get(header, "") for header in headers])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(12, min(max_length + 2, 36))

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    headers_out = {
        "Content-Disposition": 'attachment; filename="chronocampus-python-timetable.xlsx"'
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_out
    )
